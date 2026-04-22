"""
groundtruth/extract_groundtruth.py — Extract ground truth from CRA XLSX
========================================================================
Uses openpyxl (pure Python) — works on Windows, Mac, Linux.
No external CLI tools required.
"""

import os, json, re
from typing import List, Dict, Any


def _parse_xlsx_openpyxl(xlsx_path: str) -> Dict[str, str]:
    """Extract text per sheet from XLSX using openpyxl (cross-platform)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            line = "\t".join(cells)
            if line.strip():
                lines.append(line)
        sheets[sheet_name] = "\n".join(lines)
    wb.close()
    return sheets


def _extract_qa_from_sheet(sheet_name: str, text: str) -> List[Dict]:
    """Extract Q&A pairs from a single CRA sheet."""
    pairs = []
    q_pattern = re.compile(r'(Q\d+\.[\d.]+)\t(.+?)(?:\t|$)', re.MULTILINE)
    matches = list(q_pattern.finditer(text))

    for i, match in enumerate(matches):
        q_id = match.group(1).strip()
        q_text = match.group(2).strip()
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        block = text[start:end]

        answer = ""
        for pattern in [r'Answer\s*\n?\t?(.+?)(?:\n|$)', r'Response\s*\n?\t?(.+?)(?:\n|$)']:
            m = re.search(pattern, block)
            if m:
                answer = m.group(1).strip()
                break
        if not answer or answer == "Answer":
            for line in block.split("\n"):
                if "Answer" in line:
                    parts = line.split("\t")
                    for p in parts:
                        p = p.strip()
                        if p and p != "Answer" and len(p) > 2:
                            answer = p
                            break

        comments = ""
        cm = re.search(r'Comments?\s*\n?\t?(.+?)(?:\nQ\d|$)', block, re.DOTALL)
        if cm:
            comments = cm.group(1).strip().replace("\t", " ").replace("\n", " ")[:2000]

        source = ""
        sm = re.search(r'Source (?:title|URL)\s*\n?\t?(.+?)(?:\n|$)', block)
        if sm:
            source = sm.group(1).strip()

        if q_text and (answer or comments):
            difficulty = "easy"
            if any(kw in q_text.lower() for kw in ["scenario", "probability", "financial impact", "evaluated"]):
                difficulty = "hard"
            elif any(kw in q_text.lower() for kw in ["assess", "measures", "roadmap", "incentive"]):
                difficulty = "medium"

            final_answer = answer if len(answer) > 5 else (comments[:500] if comments else answer)
            pairs.append({
                "question_id": q_id, "section": sheet_name,
                "question": q_text, "answer": final_answer,
                "comments": comments[:1000], "source": source, "difficulty": difficulty,
            })
    return pairs


SUPPLEMENTARY_QA = [
    {"question_id":"S0.1","section":"0. Cover","question":"What sector is Shell Petroleum Company Limited classified under?","answer":"O&G Producer","difficulty":"easy"},
    {"question_id":"S0.2","section":"0. Cover","question":"What is the entity classification and review type?","answer":"Existing to Bank, Periodic Review, Assessment Type: Review, CRQ Version 3","difficulty":"medium"},
    {"question_id":"S6.1","section":"6. Outcome","question":"What is Shell's overall CRA score and BRAG rating?","answer":"CRA score is 57.22 with RED BRAG rating","difficulty":"easy"},
    {"question_id":"S6.2","section":"6. Outcome","question":"What are the individual section scores compared to sector averages?","answer":"Gross Physical Risk: 0.650 vs 0.650. Physical Adaptation: 0.450 vs 0.390. Gross Transition Risk: 0.197 vs 0.360. Credible Transition Plan: 0.992 vs 0.690","difficulty":"hard"},
    {"question_id":"S6.3","section":"6. Outcome","question":"Why did the BRAG rating change from AMBER to RED?","answer":"Due to more accurate approach assessing the value chain — whole supply chain including Upstream Exploration and Production was taken into consideration","difficulty":"hard"},
    {"question_id":"S7.1","section":"7. Risk Triggers","question":"What physical risk trigger was identified and what is the monitoring frequency?","answer":"Operating locations not representative. Frequency: 9 months. Track receipt of operating location address","difficulty":"medium"},
    {"question_id":"X1","section":"Cross-source","question":"What was Shell's total CFFO in 2023 and how was it distributed?","answer":"$54 billion CFFO. $23 billion shareholder returns (42% of CFFO): $8B dividends + $15B buybacks","difficulty":"hard"},
    {"question_id":"X2","section":"Cross-source","question":"What impairment risk does Shell face under IEA NZE scenario?","answer":"IG assets show $15-20B lower recoverable. Upstream $3-5B lower. Products $3-4B lower vs carrying amounts","difficulty":"hard"},
    {"question_id":"X3","section":"Cross-source","question":"What is Shell's NCI and Paris alignment approach?","answer":"NCI was 74 gCO2e/MJ in 2023, 6.3% reduction from 2016. Target 15-20% by 2030. Uses IPCC AR6 1.5C C1/C2 scenarios","difficulty":"hard"},
    {"question_id":"X4","section":"Cross-source","question":"What is Shell's R&D expenditure on decarbonisation?","answer":"Total R&D $1,287 million in 2023. Decarbonisation: $628 million (49% of total), up from 41% in 2022","difficulty":"hard"},
]


def extract_all(xlsx_path: str, output_path: str) -> Dict[str, Any]:
    """Extract all ground truth Q&A pairs from CRA XLSX."""
    print(f"    Parsing CRA: {xlsx_path}")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"CRA XLSX not found: {xlsx_path}")

    sheets = _parse_xlsx_openpyxl(xlsx_path)
    print(f"    Found {len(sheets)} sheets")

    all_qa = []
    for sheet_name, text in sheets.items():
        pairs = _extract_qa_from_sheet(sheet_name, text)
        all_qa.extend(pairs)
        print(f"      {sheet_name}: {len(pairs)} Q&A pairs")

    for sq in SUPPLEMENTARY_QA:
        all_qa.append({**sq, "comments": "", "source": "CRA Manual Analysis"})
    print(f"    + {len(SUPPLEMENTARY_QA)} supplementary questions")

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_qa, f, indent=2, ensure_ascii=False)

    print(f"    Total: {len(all_qa)} ground truth pairs → {output_path}")
    return {"total_qa": len(all_qa), "sheets": len(sheets)}
