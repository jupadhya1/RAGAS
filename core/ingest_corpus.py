"""
core/ingest_corpus.py — Document Ingestion & Chunking (v3 — All Improvements)
==============================================================================
IMPROVEMENTS:
  P0: Table-aware XLSX chunking — each Q→A→Comment block = 1 chunk
  P0: Outcome / Cover sheets kept as single chunks (decimal scores preserved)
  P0: Section header prefix on every CRA chunk for retrieval context
  NEW: Source label detection (ETS_2024, AR_2023, CDP_2023)
  NEW: Chunk metadata includes sheet, question_id, chunk_type
"""

import os, json, re, hashlib
from typing import List, Dict, Optional


# ═══════════════════════════════════════════════════════════════════════════════
# PDF EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_pdf_text(pdf_path: str) -> List[Dict]:
    """Extract text per page from PDF using pypdf (cross-platform)."""
    from pypdf import PdfReader
    reader = PdfReader(pdf_path)
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append({"page": i + 1, "text": text.strip()})
    return pages


# ═══════════════════════════════════════════════════════════════════════════════
# XLSX EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_xlsx_sheets(xlsx_path: str) -> List[Dict]:
    """Extract structured data per sheet from XLSX using openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            rows.append(cells)
        text = "\n".join(
            "\t".join(cells) for cells in rows if any(c.strip() for c in cells)
        )
        sheets.append({"sheet": sheet_name, "text": text, "rows": rows})
    wb.close()
    return sheets


# ═══════════════════════════════════════════════════════════════════════════════
# RECURSIVE TEXT CHUNKING (for PDFs)
# ═══════════════════════════════════════════════════════════════════════════════

def _chunk_text_recursive(
    text: str, chunk_size: int = 800, overlap: int = 200
) -> List[str]:
    """Recursive character splitting with sentence boundary awareness."""
    separators = ["\n\n", "\n", ". ", "? ", "! ", "; ", ", ", " "]

    def _split(text, seps, cs, ov):
        if len(text) <= cs:
            return [text] if text.strip() else []
        sep = seps[0] if seps else " "
        parts = text.split(sep)
        current = ""
        result = []
        for part in parts:
            candidate = current + sep + part if current else part
            if len(candidate) > cs and current:
                result.append(current.strip())
                overlap_text = current[-ov:] if ov > 0 else ""
                current = overlap_text + sep + part
            else:
                current = candidate
        if current.strip():
            result.append(current.strip())
        final = []
        for r in result:
            if len(r) > cs * 1.5 and len(seps) > 1:
                final.extend(_split(r, seps[1:], cs, ov))
            else:
                final.append(r)
        return final

    return _split(text, separators, chunk_size, overlap)


# ═══════════════════════════════════════════════════════════════════════════════
# P0: TABLE-AWARE XLSX CHUNKING
# ═══════════════════════════════════════════════════════════════════════════════

def _chunk_xlsx_table_aware(sheets: List[Dict]) -> List[Dict]:
    """
    P0 FIX: Table-aware chunking for CRA XLSX.

    Rules:
    1. Each Q→A→Comment block becomes ONE chunk (answer context preserved)
    2. Cover / Outcome / Risk Triggers sheets without Q-patterns → single chunk
    3. Outcome section scores (0.650, 0.992 etc.) kept in one chunk
    4. Very large blocks (>2500 chars) split at sub-question boundaries
    5. Every chunk gets a [CRA Section: X] prefix for retrieval context
    """
    chunks = []

    for sheet_data in sheets:
        sheet_name = sheet_data["sheet"]
        text = sheet_data["text"]

        q_pattern = re.compile(r"^(Q\d+\.[\d.]+)\t", re.MULTILINE)
        matches = list(q_pattern.finditer(text))

        # ── Sheets WITHOUT Q-patterns (Cover, Outcome, Risk Triggers) ──
        if not matches:
            if text.strip() and len(text.strip()) > 30:
                if len(text) <= 3000:
                    # Keep as single chunk — preserves decimal scores together
                    chunks.append({
                        "text": f"[CRA Section: {sheet_name}]\n{text}",
                        "source": "CRA_XLSX",
                        "sheet": sheet_name,
                        "chunk_type": "full_section",
                        "question_id": "",
                    })
                else:
                    # Large non-Q sheet: split with bigger chunk size
                    sub_chunks = _chunk_text_recursive(text, 1500, 300)
                    for sc in sub_chunks:
                        chunks.append({
                            "text": f"[CRA Section: {sheet_name}]\n{sc}",
                            "source": "CRA_XLSX",
                            "sheet": sheet_name,
                            "chunk_type": "section_part",
                            "question_id": "",
                        })
            continue

        # ── Header/scores BEFORE first Q ────────────────────────────
        header = text[: matches[0].start()].strip()
        if header and len(header) > 30:
            chunks.append({
                "text": f"[CRA Section: {sheet_name}]\n{header}",
                "source": "CRA_XLSX",
                "sheet": sheet_name,
                "chunk_type": "section_header",
                "question_id": "",
            })

        # ── Each Q block as ONE chunk ───────────────────────────────
        for i, match in enumerate(matches):
            start = match.start()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
            block = text[start:end].strip()

            if not block or len(block) < 20:
                continue

            q_id = match.group(1)

            if len(block) > 2500:
                # Try sub-question split first
                sub_q = re.compile(r"\n(Q\d+\.[\d.]+\.\d+)\t")
                sub_matches = list(sub_q.finditer(block))
                if sub_matches:
                    prev_end = 0
                    for j, sm in enumerate(sub_matches):
                        sub_end = (
                            sub_matches[j + 1].start()
                            if j + 1 < len(sub_matches)
                            else len(block)
                        )
                        sub_block = block[prev_end:sub_end].strip()
                        if sub_block:
                            chunks.append({
                                "text": f"[CRA {sheet_name} / {q_id}]\n{sub_block}",
                                "source": "CRA_XLSX",
                                "sheet": sheet_name,
                                "chunk_type": "qa_block",
                                "question_id": q_id,
                            })
                        prev_end = sm.start()
                else:
                    # No sub-questions — split with larger size
                    sub_chunks = _chunk_text_recursive(block, 1500, 300)
                    for sc in sub_chunks:
                        chunks.append({
                            "text": f"[CRA {sheet_name} / {q_id}]\n{sc}",
                            "source": "CRA_XLSX",
                            "sheet": sheet_name,
                            "chunk_type": "qa_block_part",
                            "question_id": q_id,
                        })
            else:
                # Normal Q block — keep intact
                chunks.append({
                    "text": f"[CRA {sheet_name} / {q_id}]\n{block}",
                    "source": "CRA_XLSX",
                    "sheet": sheet_name,
                    "chunk_type": "qa_block",
                    "question_id": q_id,
                })

        # ── Trailing content after last Q (scores, notes) ──────────
        if matches:
            last_q_end = matches[-1].end()
            # Find end of last Q block
            trailing_start = len(text)
            for i, match in enumerate(matches):
                if match == matches[-1]:
                    trailing_start = match.start()
            # The trailing after the last Q's content
            last_block_end = len(text)
            trailing = ""
            if len(matches) > 0:
                last_q_start = matches[-1].start()
                # trailing is already captured in the last Q block above
                pass

    return chunks


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _make_chunk_id(source: str, text: str) -> str:
    """Generate deterministic chunk ID."""
    h = hashlib.md5(f"{source}:{text[:100]}".encode()).hexdigest()[:12]
    return f"{source.replace(' ', '_')[:20]}_{h}"


def _detect_source_label(basename: str) -> str:
    """Map PDF filename to a clean source label."""
    bn = basename.lower()
    if "energy-transition" in bn or "ets" in bn:
        return "ETS_2024"
    elif "annual-report" in bn or "annual_report" in bn:
        return "AR_2023"
    elif "cdp" in bn:
        return "CDP_2023"
    else:
        return basename.replace(".pdf", "").replace("-", "_")[:30]


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN INGESTION
# ═══════════════════════════════════════════════════════════════════════════════

def ingest_all(
    pdf_paths: List[str],
    xlsx_path: str,
    output_dir: str,
    chunk_size: int = 800,
    chunk_overlap: int = 200,
) -> Dict:
    """Ingest all documents into a unified corpus."""
    os.makedirs(output_dir, exist_ok=True)
    corpus = []
    stats = {}

    # ─── XLSX: Table-aware Q&A chunking (P0 FIX) ─────────────
    if xlsx_path and os.path.exists(xlsx_path):
        print(f"    Ingesting XLSX: {os.path.basename(xlsx_path)} (P0: table-aware)")
        sheets = _extract_xlsx_sheets(xlsx_path)
        xlsx_chunks = _chunk_xlsx_table_aware(sheets)
        for chunk in xlsx_chunks:
            corpus.append({
                "id": _make_chunk_id("CRA_XLSX", chunk["text"]),
                "text": chunk["text"],
                "source": "CRA_XLSX",
                "source_file": os.path.basename(xlsx_path),
                "metadata": {
                    "sheet": chunk.get("sheet", ""),
                    "chunk_type": chunk.get("chunk_type", ""),
                    "question_id": chunk.get("question_id", ""),
                },
            })
        stats["CRA_XLSX"] = len(xlsx_chunks)
        print(f"      → {len(xlsx_chunks)} chunks from {len(sheets)} sheets (Q&A block-aware)")

    # ─── PDFs: Recursive chunking with page metadata ──────────
    for pdf_path in pdf_paths:
        if not os.path.exists(pdf_path):
            print(f"    ⚠️  PDF not found: {pdf_path}")
            continue

        basename = os.path.basename(pdf_path)
        source_name = _detect_source_label(basename)
        print(f"    Ingesting PDF: {basename} → {source_name}")

        pages = _extract_pdf_text(pdf_path)
        full_text = "\n\n".join(p["text"] for p in pages)
        chunks = _chunk_text_recursive(full_text, chunk_size, chunk_overlap)

        pdf_chunks = []
        for chunk_text in chunks:
            pdf_chunks.append({
                "id": _make_chunk_id(source_name, chunk_text),
                "text": chunk_text,
                "source": source_name,
                "source_file": basename,
                "metadata": {"pages": len(pages)},
            })
        corpus.extend(pdf_chunks)
        stats[source_name] = len(pdf_chunks)
        print(f"      → {len(pdf_chunks)} chunks from {len(pages)} pages")

    # Save corpus
    corpus_path = os.path.join(output_dir, "corpus.json")
    with open(corpus_path, "w", encoding="utf-8") as f:
        json.dump(corpus, f, indent=2)

    total_chars = sum(len(c["text"]) for c in corpus)
    print(f"    Total corpus: {len(corpus)} chunks, {total_chars:,} chars → {corpus_path}")
    return {"total_chunks": len(corpus), "total_chars": total_chars, "sources": stats}
